import datetime as dt
import ulazniPodaciPrimjer, postavkaUlaza, postavkaUlazaNac,rutepovrata2
import LPprogrami, clarkModul2, LPprogrami1, najnovijiDodatak8b9primjer 
import LPprogrami1Primjer,LPprogrami1aPrimjer,crtaj_rute, clarkWright

#import clarkModul1novi3, clarkModul2novi2, najnovijiDodatak8b10primjer, najnovijiDodatak8b11, Inter_routePrimjer, ulazniPodaci
from openpyxl import load_workbook
from scipy.spatial import distance
import itertools, simplejson, json
from openpyxl.styles import Font

import pulp
from collections import defaultdict
from pulp import *
global a_time,workbook3, tlong
import pickle, random
from random import choice
import numpy as np
f = open('data/hello_instance.txt','a');f.write("16 paleta ***************************************")

def print_result(tekst):
        global a_time
        b_time=dt.datetime.now()
        trajanje=(b_time-a_time).seconds
        #print (tekst + ".  " + "Vreme:" + str (trajanje)+"s")
        a_time=dt.datetime.now()
a1=dt.datetime.now()

# Maksimalno trajanje rute u satima (arbajt) i mjerna jedinica za udaljenosti (mj)
# arbajt: maksimalno dozvoljeno trajanje rute (npr. 9 sati)
# mj: 1 = metri, 1000 = kilometri
arbajt = 9
mj = 1000

# Whether to enforce the fixed-number-of-routes constraint in LP (`rutabroj`)
# Set to False to disable the constraint when calling `izracunaj_rute`
include_rutabroj = True

#vehicles = {'kombi': (5, 0.0005, 20, 1500, 1, 5), 'kamion_3.5t':(15, 0.0007, 20, 3500, 2, 7), 'kamion_7t': (20, 0.0009, 10, 7000, 3, 9)}


#wb = load_workbook(filename = 'D:\Google Drive\VOLI_konacno.xlsx')
#wb1 = load_workbook(filename = 'D:\Google Drive\Matrica_voli.xlsx')
#wb2 = load_workbook(filename = 'D:\Google Drive\Matrica_voli_time.xlsx')
""" Učitavaju se instance za tstiranje iz wb, matrica rastojanja iz wb1, matrica vremena iz wb2, a u wb3 se upisuju rezultati za svaku od instanci. """
wb = load_workbook(filename = 'data/instances_sve.xlsx')
wb1 = load_workbook(filename = 'data/hello.xlsx')
wb2 = load_workbook(filename = 'data/hello_time.xlsx')
wb3 = load_workbook('data/hello_rjesenje.xlsx')
ws3 = wb3["Sheet1"]; ws4 = wb3["Sheet2"]; ws5 = wb3["Sheet3"]; ws6 = wb3["Sheet4"] 
#instances=["GA1","GA2","GA3","GA4","GB1", "GB2","GB3","GD1"]

instances=["DE1"]
#prije pokretanja treba podesiti arbajt u najnoviji primjer 9h, to se odnosi na maksimlano trajanje rute. Prilikom pravljanje ruta se uzima to ogranicenje, ne u Lp modelu
# u najnovijem dodatku treba podesiti i mj tj mjernu jedinicu na 1000 jer su kod DL (gradska) i DE (nacionalna) razdaljine u metrima.
# U LP samo treba smaknuti ogranicenje o dozvoljenom broju ruta, koje vazi za bencmark probleme.
# za izabrani skup klijenata, problem se prvo rjesava za varijantu bez ograničenog vremena na utovaru pa za varijantu sa ograncenim. Dva puta se formira skup ruta. Pri pozivu najnovi
#Pri pozivu najnoviji dodatak je stavljeno da je 3 sata dozvoljen rad na utovaru. A vrijeme trajanje rute se mora definisati u nanovijem dodatku.
#To se radi za staro i novo stanje, i za svaki period. 

for sheet in instances:
        print()        
        print()
        oznaka=sheet[0:2]
        rute_iz_res=defaultdict(list)
        rute_iz_resui=defaultdict(list)

        a_time=dt.datetime.now()
        """ ovo sheet[0:2] sluzi da bi se od stringa koji predstavlja promjenljiva sheet uzme samo dio strinda od 0-2 koji je zapravo i naziv sheeta u excelu """
        #ws=wb[sheet]
        ws=wb[sheet[0:2]]; ws1=wb1[sheet[0:2]]; ws2=wb2[sheet[0:2]]
        """ ovo se sve čita iz podataka iz o instanci: ukupan broj čvorova koji je potrebno obići, koliko je moguće natovariti jedinica u vozilo, koji je dozvoljeni broj vozila-ruta, kolika je minimalna f-ja cilja """
        damp=sheet+".damp"; #brojCvorova=ws1.max_row
        brojCvorova=int(sheet[2:]); nosivost_cll= ws["I4"].value; br_vozila=ws["J4"].value; granica=0; benchm_broj_ruta=str(ws["J4"].value); benchm_objective=str(ws["K4"].value);strnos_cll= str(ws["I4"].value)
        """ vehicles rječnik u kome je ključ tip vozila, a vrijednosti: nosivost cll (zavisi od instance), trošak po m (samo za realni problem),
        broj vozila ovog tipa (za realni), nosivost kg (za realni), pristupnost objektu i trošak pokretanja"""
        vehicles = {'kombi': (nosivost_cll, 0.0005, 20, 1500, 1, 5)};br_koleta=0
        
        #promijeni sema zapamcena broj !!! i u instance sve kapacitet vozila                
        print("krece racunanje VRPB" + sheet)
        brk=200
        vector_opsti=[0 for x in range(brk+1)]
        for i in range(1, brk+1):
                vector_opsti[i]=[float(ws.cell(row = abs(i), column = 1).value),float(ws.cell(row = abs(i), column = 2).value)]

        # pokušaj učitavanja prethodno spremljenih šema, pada na korumpiran/nenađen pickle
        try:
                with open('data/sema_zapamcena6', 'rb') as file:
                        sve_seme = pickle.load(file)
        except (FileNotFoundError, pickle.UnpicklingError, EOFError, AttributeError, ImportError, IndexError) as e:
                print('Warning: failed to load data/sema_zapamcena6:', e)
                """# regeneriši šeme i sačuvaj ih za sljedeće pokretanje
                try:
                        sve_seme = postavkaUlaza.generisanje_sema(1)
                except Exception as g:
                        print('Error generating seme via postavkaUlaza.generisanje_sema:', g)
                        # kao krajnji fallback, generiši minimalnu strukturu koja sprječava pad
                        sve_seme = [([], {}, [], {}, [])]
                try:
                        with open('data/sema_zapamcena6', 'wb') as semf:
                                pickle.dump(sve_seme, semf)
                except Exception as h:
                        print('Warning: could not write data/sema_zapamcena6:', h)
        # ensure the instance log file handle is the expected text append handle"""
        f = open('data/hello_instance.txt','a')
        #print sve_seme        
        #sve_seme=[1]
        s=-1; sh=[1,11,2,22,3,33, 1,11,2,22,3,33, 1,11,2,22,3,33, 1,11,2,22,3,33, 1,11,2,22,3,33, 1,11,2,22,3,33, 1,11,2,22,3,33, 1,11,2,22,3,33, 1,11,2,22,3,33, 1,11,2,22,3,33, 1,11,2,22,3,33]
        ws_sema={1:ws3, 11:ws4, 2:ws3, 22:ws4, 3:ws3, 33:ws4}; ws_sema_ui={1:ws5, 11:ws6, 2:ws5, 22:ws6, 3:ws5, 33:ws6}
        brt=ws5.max_row
        for sema in sve_seme:
                rute_iz_res=defaultdict(list)
                rute_iz_resui=defaultdict(list)
                (randomklijenti_poz,kolicine_poz,randomklijenti_neg,kolicine_neg,randomklijenti_svi)=sema
                s=s+1;strs=str(sh[s]);strws=ws_sema[sh[s]];strws_ui=ws_sema_ui[sh[s]]
                if sh[s]<10:
                        brt=brt+1; #broj tesiranih podinsaci kao kod. povecava se kad se sema promijeni podinstanca
                        print("broj testiranih podinstanci:", brt)
                        f.write(str(brt))
                sheet=str(nosivost_cll)+oznaka+str(brt)
                f.write(str(brt)+str(sh[s])+"\n")
                f.write("randomklijenti_poz="+ str(randomklijenti_poz)+ "\n"); f.write("kolicine_poz="+ str(kolicine_poz)+ "\n"); f.write("randomklijenti_neg="+ str(randomklijenti_neg)+ "\n");f.write("kolicine_neg="+ str(kolicine_neg)+ "\n");f.write("randomklijenti_svi="+ str(randomklijenti_svi)+ "\n"); 
                        
                #randomklijenti_poz=[124, 90, 115, 85, 81, 68, 40, 130, 116, 59, 7, 191, 187, 46, 93, 23, 31, 188, 151, 19, 155, 21, 126, 48, 97, 141, 45, 200, 77, 75, 41, 125, 174, 70, 28, 54, 190, 194, 102, 170, 96, 44, 185, 147, 179, 178, 47, 162, 110, 104, 121, 171, 8, 199, 38, 134, 181, 149, 84, 153, 18, 182, 152, 103, 114, 122, 25, 20, 123, 60]
                #kolicine_poz={1: 0, 130: 1, 134: 3, 7: 1, 8: 3, 23: 1, 141: 2, 18: 3, 19: 1, 20: 3, 21: 2, 151: 1, 152: 3, 25: 3, 155: 2, 28: 2, 31: 1, 162: 2, 38: 3, 40: 1, 41: 2, 147: 2, 171: 3, 44: 2, 45: 2, 46: 1, 47: 2, 48: 2, 178: 2, 179: 2, 181: 3, 54: 2, 185: 2, 187: 1, 188: 1, 190: 2, 191: 1, 194: 2, 68: 1, 182: 3, 70: 2, 199: 3, 200: 2, 75: 2, 77: 2, 81: 1, 84: 3, 85: 1, 90: 1, 93: 1, 96: 2, 97: 2, 59: 1, 102: 2, 103: 3, 104: 2, 60: 3, 125: 2, 110: 2, 174: 2, 114: 3, 115: 1, 116: 1, 153: 3, 121: 3, 122: 3, 123: 3, 124: 1, 170: 2, 126: 2, 149: 3}
                #randomklijenti_neg=[-124, -125, -31, -130, -28, -153, -93, -8, -45, -171, -41, -48, -38, -18, -170]
                #kolicine_neg={-31: 1, -125: 1, -124: 1, -170: 2, -153: 1, -18: 2, -48: 1, -93: 1, -171: 1, -45: 1, -41: 1, -8: 1, -28: 1, -38: 2, -130: 1}
                #randomklijenti_svi=[124, 90, 115, 85, 81, 68, 40, 130, 116, 59, 7, 191, 187, 46, 93, 23, 31, 188, 151, 19, 155, 21, 126, 48, 97, 141, 45, 200, 77, 75, 41, 125, 174, 70, 28, 54, 190, 194, 102, 170, 96, 44, 185, 147, 179, 178, 47, 162, 110, 104, 121, 171, 8, 199, 38, 134, 181, 149, 84, 153, 18, 182, 152, 103, 114, 122, 25, 20, 123, 60, -124, -125, -31, -130, -28, -153, -93, -8, -45, -171, -41, -48, -38, -18, -170]                

                (dsv,tsv, zao)=ulazniPodaciPrimjer.izracunaj_matricu_opstu (ws, ws1, ws2,brk, randomklijenti_poz, kolicine_poz, randomklijenti_neg, kolicine_neg)
                #break
                #print ("odredjuje se matrica za sve klijente i matrica samo za klijente do kojih se razvozi")
                """računa se matrica najkracih rastojanja izmedju svih cvorova i isporukama i sa povratom"""
                #(dsv,tsv, zao)=ulazniPodaci.izracunaj_matricu_opstu(ws, ws1, ws2, brojCvorova, -1000000)
                #vreme_rada=nosivost_cll+((brojCvorova-1)/(br_koleta/nosivost_cll))*100+10000000
                vreme_rada=9
                print("vreme rada:",vreme_rada) 

                """ računanje matrice rastojanja samo za klijente za koje se roba isporučuje + depo (zato stoji nula u argumentu), gdje cvorovi dobijaju nove brojeve koji se smjestaju u listu fields sa prvim clanom skladistem (1) """
                (d,durations, osobine,broj_klijenata, vector, br_koleta, uk_masa)=ulazniPodaciPrimjer.izracunaj_matricu(ws, ws1, ws2, brk, 0, randomklijenti_poz, kolicine_poz)
                
                
                #print ("sad se odredjuju gigantske razvozne")
                b_time=dt.datetime.now()
                """ end predstavlja max vrijednost za taboo"""
                end=max(5,round((broj_klijenata-1)/4.5))
                #end=min(15,round(br_klijenata_poz/4.5))
              
                """ ovaj modul isporučujw rute razvoza. to su različite kombinacije covrova koji se nalaze u pozitivnoj matrici.
                Najmanje moze biti jedan, a najviše za svaku taboo vrijednos po jedno rješenje (plus obrnuto) i plus pocetno za klark. Svako od rješenja se kroz pripremu za dijsktru cjepkaju na pojedinacne rute"""

                """ako se uzimaju velike rute razvoza kao rjesenje od TSP sa povratkom u bazu"""
                #rute_razvoza=clarkModul1novi3.racunaj(d,vector, broj_klijenata, end)
                rute_razvoza=defaultdict(list)
                
                """ovo je za slucaj kada se uzimaju rjesenja bez povrata u bazu. moze da ukljuci i prethodne rute ako se ostave ukljucene kao argument za sledece."""
                #rute_razvoza=clarkModul2novi2.racunaj(d,vector, broj_klijenata, end, rute_razvoza)
                rute_razvoza=clarkWright.racunaj(d,rute_razvoza)
                
                c_time=dt.datetime.now(); calc_time=(c_time-b_time).seconds 
                print("trajanje racunanje gigantskih ruta razvoza" + str(calc_time))
                
                #print ("sad se rade gigantske povratne")
                """Računa se matrica rastojanja između klijenata sa povratom. U ovom slucaju fields se zavrsavaju sa skladistem, tj cvorom 1.
                Zato ce morati imati i svoj Clarck modul jer ce za razliku od razvoza poslenji cvor biti 1 odmah clark nije isti"""
                (d1,durations1, osobine1,broj_klijenata1, vector1, br_koleta1, uk_masa1)=ulazniPodaciPrimjer.izracunaj_matricu_neg(ws, ws1, ws2, brk, 0, randomklijenti_neg, kolicine_neg)
                fields=list(range(2,broj_klijenata1+1))+[1]
                d_time=dt.datetime.now()
                """lista rute_povrata prestavlja skup gigantskih ruta povrata u kojima je svaki od čvorova sa povratom prvi bar jednoj.
                Neki cvor sa povratom može imati i nekoliko gigantskih ruta. Svaka od njih uključuje sve ostale čvorove sa povratom i završava u čvoru 1"""
                rute_povrata=rutepovrata2.lista_ruta(d1,vector1, broj_klijenata1, fields, broj_klijenata1)
                rute_povrata_sortirane=sorted(rute_povrata.items(), key=lambda x: x[1])
                #print "rute_povrata_sortirane", rute_povrata_sortirane
                e_time=dt.datetime.now(); calc_time=(e_time-d_time).seconds
                print("trajanje racunanje gigantskih povratnih" + str(calc_time))
                
                f_time=dt.datetime.now()
                rute_razvoza_sortirane=sorted(rute_razvoza.items(), key=lambda x: x[1])
                #moglo bi ovdje da se stavi for v in vehicles i da se radi za različite nosivosti odavde na dolje
                g_time=dt.datetime.now(); calc_time=(g_time-f_time).seconds
                (routes2,skup_povratnih_svih)=najnovijiDodatak8b9primjer.spoji_rute (rute_razvoza_sortirane, rute_povrata, vehicles, osobine,osobine1, dsv,zao, vreme_rada,tsv, arbajt, mj)
                #routes3=najnovijiDodatak8b10primjer.spoji_rute (rute_razvoza_sortirane[0:1], rute_povrata_sortirane[0:1], vehicles, osobine,osobine1, dsv,zao,vreme_rada,tsv)
                #routes3=najnovijiDodatak8b11.spoji_rute (rute_razvoza_sortirane[0:2], rute_povrata_sortirane[0:2], vehicles, osobine,osobine1, dsv,zao,skup_povratnih_svih)

                """ Ovdje se dodaju u skup za izbor konacnog rjesenja sve rute koje su dobijene spajanjem"""
                rute_iz_res.update(routes2)
                #rute_iz_res.update(routes3)
                
                #print "cjepkanje povratnih i kalemljenje na razvozne" + str(calc_time)
                                        
                print("ukupno broj mogucih ruta:")
                ukupno_mogucih_ruta=len(rute_iz_res)
                print(ukupno_mogucih_ruta)                

                """Ovdje se bira skup najboljih tako da svaki cvor bude samo jednom ukljucen. Jedino sto moze biti problem je da se rijesenje ne moze naci za zadati broj ruta.
                Ovdje ne treba prebacivati u stare brojeve jer je vec sve vraceno pa je zato False"""
                "ako se slucajno zbog ogranicenja za broj ruta, ne dobije rjesenje koje je izvodljvio, broj dozvoljenih ruta-vozila se povećavaja dok se ne dobije optimalno rjesenje"
                #pickle.dump(rute_iz_res, open(damp, 'wb'))
                #rute_iz_res = pickle.load(open(damp, 'rb'))
                status="nemoguce"
                #while (not (status=="Optimal" or status== "Stopped on time limit")):
                while not status=="Optimal":
                        (frute_iz_res, frute_vozila, fbroj_ruta, fvreme_rute, objective1, status)=LPprogrami1Primjer.izracunaj_rute(randomklijenti_svi, rute_iz_res, vehicles, osobine, False, br_vozila, include_rutabroj)
                        br_vozila+=1
                        print(status)
                        print(br_vozila)
                        #print array_1d

              
                h_time=dt.datetime.now(); calc_time=(h_time-g_time).seconds
                print("racubanje rjesenja" + str(calc_time))
                LPprogrami1.printaj_rute(frute_iz_res, zao, sheet, "VRPB"+strs, frute_vozila, fbroj_ruta, strws, objective1, status, benchm_broj_ruta,benchm_objective, ukupno_mogucih_ruta, broj_klijenata-1, broj_klijenata1-1, strnos_cll, 1, ukupno_mogucih_ruta,calc_time)

                benchm_objective=str(objective1); benchm_broj_ruta=str(fbroj_ruta["kombi"])
                #(frute_iz_res2, frute_vozila, objective)=Inter_routePrimjer.testiraj(frute_iz_res, vehicles, dsv, zao, vreme_rada)
                #r_time=dt.datetime.now(); calc_time=(r_time-g_time).seconds
                #print "ukupno sekundi" + str(calc_time)
                
                #LPprogrami1.printaj_rute(frute_iz_res2, zao, sheet, "VRPB", frute_vozila, fbroj_ruta, ws3, objective, status, benchm_broj_ruta,benchm_objective, ukupno_mogucih_ruta, broj_klijenata-1, broj_klijenata1-1, strnos_cll, 1, ukupno_mogucih_ruta,calc_time)
                # ne treba LPprogrami1.izracunaj_vozila(rute_iz_res1, rute_vozila, broj_ruta, vreme_rute, vehicles)
                wb3.save('data/hello_rjesenje.xlsx')
                crtaj_rute.crtaj(vector_opsti, frute_iz_res)
                
                #ZA DIO SA OGRANIČENJEM TRAJANJA UTOVARA I ISTOVARA
                print()
                print(" Sada dio za LP sa ograničenjima")
                g_time=dt.datetime.now()
                (routes2ui,skup_povratnih_svih)=najnovijiDodatak8b9primjer.spoji_rute (rute_razvoza_sortirane, rute_povrata, vehicles, osobine,osobine1, dsv,zao, 3,tsv, arbajt, mj)
                rute_iz_resui.update(routes2ui)
                ukupno_mogucih_ruta=len(rute_iz_resui)
                print("ukupno broj mogucih ruta:" + str(ukupno_mogucih_ruta))

                #rute_iz_res = pickle.load(open(damp, 'rb'))
                status="nemoguce"
                while (not (status=="Optimal" or status== "Stopped on time limit")):
                #while not status=="Optimal":
                        (frute_iz_res, frute_vozila, fbroj_ruta, fvreme_rute, objective, status)=LPprogrami1aPrimjer.izracunaj_rute(randomklijenti_svi, rute_iz_resui, vehicles, osobine, False, br_vozila)
                        br_vozila+=1
                        print(status)
                        print(br_vozila)
                        
                h_time=dt.datetime.now(); calc_time=(h_time-g_time).seconds
                print("racubanje rjesenja" + str(calc_time))
                LPprogrami1.printaj_rute(frute_iz_res, zao, sheet, "VRPBsaUI"+strs, frute_vozila, fbroj_ruta, strws_ui, objective, status, benchm_broj_ruta, str(objective1), ukupno_mogucih_ruta, broj_klijenata-1, broj_klijenata1-1, strnos_cll, 1, ukupno_mogucih_ruta,calc_time)
                #(frute_iz_res2, frute_vozila, objective)=Inter_routePrimjer.testiraj(frute_iz_res, vehicles, dsv, zao, vreme_rada)
                #r_time=dt.datetime.now(); calc_time=(r_time-g_time).seconds
                #print "ukupno sekundi" + str(calc_time)
                
                #LPprogrami1.printaj_rute(frute_iz_res2, zao, sheet, "VRPB", frute_vozila, fbroj_ruta, ws3, objective, status, benchm_broj_ruta,benchm_objective, ukupno_mogucih_ruta, broj_klijenata-1, broj_klijenata1-1, strnos_cll, 1, ukupno_mogucih_ruta,calc_time)
                # ne treba LPprogrami1.izracunaj_vozila(rute_iz_res1, rute_vozila, broj_ruta, vreme_rute, vehicles)
                wb3.save('data/hello_rjesenje.xlsx')
f.close()


