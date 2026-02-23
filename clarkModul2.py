def racunaj(d,vector, brojCvorova,end_raz,rute_razvoza):
   
 ################
    from openpyxl import load_workbook
    #wb = load_workbook(filename = 'D:\Google Drive\Voli_konacno.xlsx');ws = wb['Sheet2']
    import datetime as dt
    global t
    #ucitavanje potraznje p za clark wrightov algoritam
    p=[0 for x in range(brojCvorova+1)]
    from collections import defaultdict
    #wb = load_workbook(filename = 'E:\euclid_podaci.xlsx')
    #ws = wb[sheet]
    #for i in range(1,brojCvorova+1):
            #p[j]=ws.cell(row = i, column = 3).value
    #rute_razvoza=defaultdict(list)
    print "broj gigantski ruta razvoza sa depom kao ciljem: " + str(len(rute_razvoza))

    class Usteda:
        def __init__ (self, cvori, cvorj,vrednost):
            self.cvori=cvori
            self.cvorj=cvorj
            self.vrednost=vrednost
            
        def __repr__(self):
            return '{}: {} {} {}'.format(self.__class__.__name__,
                                      self.cvori,
                                      self.cvorj,
                                      self.vrednost)
        def __cmp__(self, other):
            if hasattr(other, 'vrednost'):
                c=self.vrednost-other.vrednost    
                return int(c)

        def __cmp_float__(self, other):
            if hasattr(other, 'vrednost'):
                return self.vrednost.__cmp__(other.vrednost)
            
    class Route:
        def __init__ (self, rb, fields, duzina):
            self.rb=rb
            self.fields=fields
            self.duzina=duzina
            
        def __repr__(self):
            return '{}: {} {}'.format(self.__class__.__name__,
                                      self.rb,
                                      self.fields)
        #def __cmp_float__(self, other):
            #if hasattr(other, 'duzina'):
                #return self.duzina.__cmp__(other.duzina)
    class Tabo:
        def __init__ (self, i, k, value,frequency):
            self.i=i
            self.k=k
            self.value=value

    ustede=[] 
    for i in range (2,brojCvorova+1):
         for j in range(i+1,brojCvorova+1):
             vrednost=d[1][j] - d[i][j]
             ustede.append (Usteda (i,j,vrednost))
             
    ustede.sort(reverse=True)

    cvorovi=[]
    for i in range (1,brojCvorova+1):
            cvorovi.append(i)

    def printaj():
            print " ".join(["{:3d}".format(x) for x in cvorovi])
            print " ".join(["{:3d}".format(x) for x in izlazniCvorovi])
            print " ".join(["{:3d}".format(x) for x in ulazniCvorovi])
            print ("posjeceni cvorovi")
            print posjeceniCvorovi

    def sredi1(vezivniCvor):
         index1=0
         while index1 < len(r):
            if ustede[index1].cvori==vezivniCvor or ustede[index1].cvorj==vezivniCvor :
                del ustede[index1]
                r.pop()
            else:index1=index1+1

    def sredi (vanjskiCvor):
            global suma
            suma=suma+p[vanjskiCvor]
            del ustede[index]
            r.pop()
            posjeceniCvorovi.append (vanjskiCvor)
            index1=0
            for index1 in r: 
                if ustede[index1].cvori==vanjskiCvor:
                   for i in range (0, len(posjeceniCvorovi)):
                     if  ustede[index1].cvorj==posjeceniCvorovi[i]:
                        del ustede[index1]
                        r.pop()
                        break
                elif ustede[index1].cvorj==vanjskiCvor:
                  for i in range (0, len(posjeceniCvorovi)):
                     if ustede[index1].cvori==posjeceniCvorovi[i] :
                        del ustede[index1]
                        r.pop()
                        break
                    
    def izracunaj_duzinu (fields):
            zbir=0
            # ako imamo niz 5,6,7,8,3 len je 5. cetvrti clan niza je broj 3. ako je len-1 i uzim 3 element za poslednju vrednost a to je 8. 
            # stavljeno -2 umjesto -1 da se ne bi racunao povratak u bazy. tako da ukupna duzina rute ne uzima u obzir put do baze iako se pise..
            #   ..da se zavrsava u bazi tj. u 1.
            for i in range(0,len(fields)-2):
                    zbir=zbir+d[fields[i]][fields[i+1]]
            return zbir
            
    def izracunaj_2optduzinu (fields, duzina, i, k):      
            b=-d[fields[i-1]][fields[i]]-d[fields[k]][fields[k+1]]+d[fields[i-1]][fields[k]]+d[fields[i]][fields[k+1]]
            #b=-d[fields[i-1]][fields[i]]-d[fields[k]][fields[k+1]]+d[fields[i-1]][fields[k+1]]+d[fields[k]][fields[i]]
            z=duzina+b
            return z
            
    def dva_opt (fields,duzina):
            global solution,currentbestRoute
            #print solution
            #currentbestRoute=solution
            najkraca_dosad=duzina
            provera=False
            brojrute=-1;iter=0
            #len(fields)-1, -1 je da se ne bi dirala jedinica na kraju
            for i in xrange(1, len(fields)-1):
                    for k in xrange(i + 1, len(fields)-1):
                            brrute = brojrute+1
                            tos=twoOptSwap(fields, i, k)
                            dvaoptduzina=izracunaj_duzinu(tos)
                            if round (dvaoptduzina,2)<round(najkraca_dosad,2):
                                    #print ("Bio sam tuuuuuuu")
                                    #print najkraca_dosad
                                    #print dvaoptduzina
                                    currentbestRoute=Route(brojrute,tos, dvaoptduzina)
                                    #print currentbestRoute
                                    najkraca_dosad=dvaoptduzina
                                    provera=True
            if provera==True:
                    iter=iter+1
                    #print ("novo")
                    #print najkraca_dosad
                    #print ("iz 2opt:" + str(izracunaj_duzinu (currentbestRoute.fields)))
                    dva_opt (currentbestRoute.fields,najkraca_dosad)


            try:
                    #print currentbestRoute
                    solution=currentbestRoute
                    #print ("Bio sam tuuuuuuu")
            except:
                    print ("Bio sam kod excepttuuuuuuu")
                    solution=solution
            #return solution


            
    t =  [[0 for x in range(brojCvorova+1)] for y in range(brojCvorova+1)] #vrijednost tabua
    #tlong =  [[0 for x in range(brojCvorova+1)] for y in range(brojCvorova+1)]
    tabu=[0 for x in range(brojCvorova+1)]
    currentbestRoute_set=defaultdict(list)

    def tabu_deo (fields, duzina, zadato, iterator, t1, t2, rute_razvoza):
            global solution,currentbestRoute,t
            najkraca_dosad=zadato
            provera=False
            brojrute=-1;iterator=iterator
            #len(fields)-1, -1 je da se ne bi dirala jedinica na kraju niza. Dakle 1, len(fields)-1) znaci da se ne diraju prvi i zadnji element u nizu, a to su jedinice.
            #if t1>2:
                #print t1, rute_razvoza
            for i in xrange(1, len(fields)-1):
                    for k in xrange(i + 1, len(fields)-1):
                        brojrute = brojrute+1
                        tos=twoOptSwap(fields, i, k)
                        #print tos
                        tos5=tos[:-1]
                        """ Ako je rjesenje na koje se naislo vec za pretodni taboo izabrano kao najbolje ne uzima se kao sledece u pretrazi"""
                        if tuple(tos5) not in rute_razvoza:
                                duz=izracunaj_duzinu(tos)
                                #print tos, duz, solution.duzina
                                if round(duz,2)<round(najkraca_dosad,2) and round(duz,2)<round(solution.duzina,2):
                                    #print t1, tos, duz, solution.duzina, solution.rb
                                    #print t1, solution.duzina, duz, "osim", iterator
                                    #print ("duz:"+str(duz)+" najkraca:"+str (najkraca_dosad))
                                    currentbestRoute=Route(brojrute,tos, duz)
                                    solution=currentbestRoute
                                    najkraca_dosad=currentbestRoute.duzina
                                    provera=True
                                    inaj=fields[i]
                                    knaj=fields[k]
                                    #t[knaj][inaj]=t1;t[inaj][knaj]=t2
                                    #tlong[knaj][inaj]+=1;tlong[inaj][knaj]+=1
                                    tabu[inaj]=t1; tabu[knaj]=t2
                                #""" Ako se potrebno ponovo zameniti mjesta čvorovima na kojima je taboo prošao, moguče je uz povećanje rastojanja
                                #za vrijednos tlonga """
                                                               
                                elif round(duz,2)+tlong[fields[i]][fields[k]]<round(najkraca_dosad,2) and t[fields[i]][fields[k]]==0:
                                #elif round(duz,2)<round(najkraca_dosad,2) and tabu[fields[i]]==0 and tabu[fields[k]]==0:
                                #and tlong[fields[i]][fields[k]]<3:
                                    
                                    #tos6=tos
                                    #if tuple (tos6) not in currentbestRoute_set:
                                    #print ("222duz:"+str(duz)+" najkraca:"+str (najkraca_dosad))
                                    #print ("i="+str(i)+" k="+str(k))
                                    #print ("t[fields[i]][fields[k]]="+str(t[fields[i]][fields[k]]))
                                    #print ("fields_i="+str(fields[i])+" fields_k="+str(fields[k]))
                                    inaj=fields[i]
                                    knaj=fields[k]
                                    currentbestRoute=Route(brojrute,tos, duz)
                                    najkraca_dosad=currentbestRoute.duzina
                                    provera=True
                                    #print ("Bio sam u taboo2")
                                    #print currentbestRoute.duzina
                                    #tlong[knaj][inaj]+=1;tlong[inaj][knaj]+=1
                                    tabu[inaj]=t1; tabu[knaj]=t2
                                    #print tabu[inaj], tabu[knaj]
                               
                        #else:
                            #print ("ne moze ima u rute razvoza")

                                                                                                                                                                                   
            if provera==True:
                    tlong[knaj][inaj]+=100;tlong[inaj][knaj]+=100
                    t[knaj][inaj]=t1 ;t[inaj][knaj]=t2
                    #for i in fields:
                            #for k in fields:
                                #if t[i][k]>0:
                                    #t[i][k]=t[i][k]-1
                                  
                    c=0
                    for i in xrange(1, len(fields)):
                            for k in xrange(i + 1, len(fields)):
                            #for k in xrange(1, len(fields)):
                                  if t[fields[i]][fields[k]]>0:
                                            #print fields[i],fields[k],t[fields[i]][fields[k]]
                                            #print fields[k],fields[i],t[fields[k]][fields[i]]
                                            c+=1
                                            t[fields[i]][fields[k]]=t[fields[i]][fields[k]]-1
                                            #t[fields[k]][fields[i]]=t[fields[k]][fields[i]]-1
                    #print iterator, c

                    for i in xrange(1, len(fields)):
                        if tabu[i]>0:
                            tabu[i]=tabu[i]-1
                        
                    #iterator=iterator+1
                    #print ("iz taboo solution:"+str(solution.duzina))
                    #print izracunaj_duzinu (solution.fields)
                    #print solution.fields
                    #print iterator, t1,t2
                    #print currentbestRoute.fields
                    tabu_deo (currentbestRoute.fields,currentbestRoute.duzina, currentbestRoute.duzina, iterator, t1, t2, rute_razvoza)
            elif iterator <3:
                    iterator=iterator+1
                    #print iterator
                    #print ("else")
                    #print currentbestRoute.fields
                    #currentbestRoute_set[tuple(currentbestRoute.fields)]
                    #print ("currentbestRoute_set")
                    #print currentbestRoute_set
                    #t =  [[0 for x in range(brojCvorova+1)] for y in range(brojCvorova+1)]
                    #tlong =  [[0 for x in range(brojCvorova+1)] for y in range(brojCvorova+1)]
                    tabu_deo (currentbestRoute.fields,currentbestRoute.duzina, 50000000000000,iterator, t1, t2, rute_razvoza)
                    #sad mora uci u pretragu,jer ce sigurno ruta biti kraca od 5000000000
            #else:
                    #print fields
                    #for i in fields:
                            #for k in fields:
                                  #if tlong[i][k]>0:
                                            #print i,k,tlong[i][k]

    def twoOptSwap(fields, i, k):
        newFields = fields[:]
        newFields[i:k+1] = fields[k:i-1:-1]
        return newFields


    def clark():
     global novaRuta, suma,posjeceniCvorovi,fields,izlazniCvorovi,ulazniCvorovi,index,r,brojrute,listaruta, solution, currentbestRoute
     #inicijalizacija promenljivih i lista. Stavljeno je da su javni, jer se proseleduju u f-je sred i sredi1 koje ih i menjaju
     novaRuta=False
     suma=0
     posjeceniCvorovi=[]
     fields=[0 for x in range(brojCvorova+1)]
     izlazniCvorovi = [0 for x in range(brojCvorova)]
     ulazniCvorovi = [0 for x in range(brojCvorova)]
     index=0
     r = range(0, len(ustede))#broj usteda
     brojrute=-1
     listaruta=[]
     while index<len(r):
            if novaRuta==True and suma<10:
                    
               if ustede[index].cvorj==start:
                  vezivniCvor=start     
                  izlazniCvorovi[ustede[index].cvori-1]=ustede[index].cvorj
                  ulazniCvorovi[start-1]=ustede[index].cvori
                  start=ustede[index].cvori
                  sredi (start)
                  sredi1(vezivniCvor)
                  index=0
            
               elif ustede[index].cvori==start:
                  vezivniCvor=start 
                  izlazniCvorovi[ustede[index].cvorj-1]=ustede[index].cvori
                  ulazniCvorovi[start-1]=ustede[index].cvorj
                  start=ustede[index].cvorj
                  sredi(start)
                  sredi1(vezivniCvor)
                  index=0
                 
               elif ustede[index].cvori==kraj:
                  vezivniCvor=kraj 
                  izlazniCvorovi[kraj-1]=ustede[index].cvorj
                  ulazniCvorovi[ustede[index].cvorj-1]=ustede[index].cvori
                  kraj=ustede[index].cvorj
                  sredi(kraj)
                  sredi1(vezivniCvor)
                  index=0
            
               elif ustede[index].cvorj==kraj:
                  vezivniCvor=kraj
                  izlazniCvorovi[kraj-1]=ustede[index].cvori
                  ulazniCvorovi[ustede[index].cvori-1]=ustede[index].cvorj
                  kraj=ustede[index].cvori
                  sredi(kraj)
                  sredi1(vezivniCvor)
                  index=0

               else:
                       index=index+1

            elif novaRuta==False:
               suma=suma+p[ustede[index].cvori]+p[ustede[index].cvorj]
               start=ustede[index].cvori
               kraj=ustede[index].cvorj
               izlazniCvorovi[ustede[index].cvori-1]=ustede[index].cvorj
               ulazniCvorovi[ustede[index].cvorj-1]=ustede[index].cvori
               novaRuta=True
               posjeceniCvorovi.append (start)
               posjeceniCvorovi.append (kraj)
               del ustede[index]
               r.pop()
               
            #kreiranje rute iz niza posjeceniCvorovi i njeno printanje.
            #Ruta se kreira ukoliko vise nema cvorova za obilazak ili ukoliko je dostignut neki maksimum (ovde je to 10 necega u vozilu, pri cemu je na svakoj lokacija jedna jedinica-ali je gore stavljeno da je taj niz sastavljen od nula)           
            if index==len(r) or suma==10 :
               duzina=0     
               brojrute=brojrute+1
               n2=dt.datetime.now()                        
               fields[0]=1
               fields[1]=start
               i=1
      
               while i< len(posjeceniCvorovi):
                       fields[i+1]=izlazniCvorovi[fields[i]-1]
                       i=i+1
               fields[len(posjeceniCvorovi)+1]=1     
               duzina=izracunaj_duzinu(fields)
          
               solution=Route(brojrute,fields,duzina)
               currentbestRoute=solution
               sredi1 (start)
               sredi1 (kraj)
       
               listaruta.append(solution)
               
               
               novaRuta=False
               suma=0
               posjeceniCvorovi=[]
               fields=[0 for x in range(brojCvorova+1)]
               izlazniCvorovi = [0 for x in range(brojCvorova)]
               ulazniCvorovi = [0 for x in range(brojCvorova)]
               index=0

    clark()

    #print ("clark rjesenje")    
    #print solution.duzina
    #print solution.fields


    #dva_opt (solution.fields,solution.duzina)
    #print ("2opt rjesenje")
    #print solution.duzina
    #print solution.fields
    # ovdje bi trebalo zapamtiti od clarck rjesenja solution pa dolje za svako novi brojac krenuti od njega ili sve solutione stavljatu listu ili u ovaj tuple koji bi bio javan. A i clark treba da udje u rjesnje

    solution_clark=solution
    sf=solution.fields[:-1]
    #print "clarkModul2 ruta razvoza bez cilja da se vraca u depo"
    #Na kraju je dodat samo. Misim da u ukupnu duzinu ne ulazi rasotjanje od zadnjeg cvora do depoa
    #print solution.fields, solution.duzina
    rute_razvoza[tuple(sf)]


    brojac_raz=end_raz-3
    

    while not brojac_raz>end_raz:

        try:
                ##print
                
                t =  [[0 for x in range(brojCvorova+1)] for y in range(brojCvorova+1)]
                tlong =  [[0 for x in range(brojCvorova+1)] for y in range(brojCvorova+1)]
                #solution.duzina=500000000000000000
                tabu_deo (solution.fields,solution.duzina, 500000000000000000,0,brojac_raz,brojac_raz, rute_razvoza)
                ##print "novi taboooooo:", brojac_raz,solution.duzina
        except Exception as e:
                print(e)
                print ("except")

        #print (brojac_raz, "taboo rjesenje:" + str(solution.duzina))
        #print solution.duzina
        print brojac_raz, solution.fields, solution.duzina
        #print tlong

        sfr=[solution.fields]
        #ovdje solution.fields pocinje i zavrsava sa 1, pa zato i kad se okrene moze da se oduzme zadnji clan tj.1
        # kad se stavi [:-1] zadrzava se lista (solution.fields)-ne mijanja se
        # ovo [::-1] znaci obrni a da sf_reverse ostane isto poslije, tj da samo sf_r dobije obrnut redoslijed.
        sf=solution.fields[:-1]
        sf_reverse=solution.fields [1:]
        sf_r=sf_reverse[::-1]
    
        #listaRuta.append(sf)
        # i sf i sf_r su sa 1 na pocetku 
        rute_razvoza[tuple(sf)]
        rute_razvoza[tuple(sf_r)]
        brojac_raz+=1
        global solution
        solution=solution_clark


        
    #print    
    #print  rute_razvoza

    print "ukupan broj gigantskih ruta razvoza: " + str(len(rute_razvoza))
      
    return (rute_razvoza)


"""
    f = open('D:\Google Drive\sf', 'w')
    simplejson.dump(sf, f)
    f.close()
"""




    
            
             
