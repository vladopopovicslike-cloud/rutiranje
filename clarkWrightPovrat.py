# -*- coding: cp1252 -*-
def racunaj(d,vector, brojCvorova, fields1, prvic, end_raz, rute_razvoza):
   
 ################
    from openpyxl import load_workbook
    #wb = load_workbook(filename = 'D:\Google Drive\Voli_konacno.xlsx');ws = wb['Sheet2']
    import datetime as dt
    from collections import defaultdict
    global t
    #ucitavanje potraznje p za clark wrightov algoritam
   
    #wb = load_workbook(filename = 'E:\euclid_podaci.xlsx')
    #ws = wb[sheet]

    rute_razvoza1=defaultdict(list)
    rute_razvoza=rute_razvoza
    #rute_razvoza1=rute_razvoza
    #print "FSFSFSFSF", rute_razvoza1
            
    class Usteda:
        def __init__ (self, cvori, cvorj,vrednost):
            self.cvori=cvori
            self.cvorj=cvorj
            self.vrednost=vrednost
            
        def __repr__(self):
            return '{}: {} {} {}'.format(self.__class__.__name__,
                                      self.cvori,
                                      self.cvorj,
                                      self.vrednost)
        def __lt__(self, other):
            if hasattr(other, 'vrednost'):
                return self.vrednost > other.vrednost
            
    class Route:
        def __init__ (self, rb, fields, duzina):
            self.rb=rb
            self.fields=fields
            self.duzina=duzina
            
        def __repr__(self):
            return '{}: {} {}'.format(self.__class__.__name__,
                                      self.rb,
                                      self.fields)
        #def __cmp_float__(self, other):
            #if hasattr(other, 'duzina'):
                #return self.duzina.__cmp__(other.duzina)
    class Tabo:
        def __init__ (self, i, k, value,frequency):
            self.i=i
            self.k=k
            self.value=value

    ustede=[] 

    red_matrice=len(d) #"red matrice"
    broj_klijenata=len(d)-2 #broj klijenata. -2 je jer ne ulaze nula i 1
    #mora se ici sa 0 u matrici jer poslije priliko 2opt postoji 0 kao clan
    #print "red matrice",red_matrice
    ustede=[] 
    for i in range (2, red_matrice):
         for j in range(2,red_matrice):
             if i!=j and i!=prvic and j!=prvic:
                 vrednost=d[i][1]+d[prvic][j] - d[i][j]
                 ustede.append (Usteda (i,j,vrednost))
                         
    ustede.sort(reverse=True)
    #print ustede
    
    def izracunaj_duzinu (fields):
        zbir=0
        # ako imamo niz 5,6,7,8,3 len je 5. cetvrti clan niza je broj 3. ako je len-1 i uzim 3 element za poslednju vrednost a to je 8. 
        # stavljeno -2 umjesto -1 da se ne bi racunao povratak u bazy. tako da ukupna duzina rute ne uzima u obzir put do baze iako se pise..
        #   ..da se zavrsava u bazi tj. u 1.
        for i in range(0,len(fields)-1):
                zbir=zbir+d[fields[i]][fields[i+1]]
        return zbir


    cw_ruta=[]
    cw_ruta.append(ustede[0].cvori);cw_ruta.append(ustede[0].cvorj)
    #-1 zato sto prvi red u matrici ide 0,1,2.. a ovdje se 2 uzima kao prvi cvor rute pa je duzina rute na kraju manja za 1 i zato mora i broj klijenata da se smanji za 1
    while len (cw_ruta)<broj_klijenata-1:
        #print "idem iznova"
        for u in ustede:
            #print "usteda", u
            if u.cvori==cw_ruta[-1]:
                if u.cvorj not in cw_ruta:
                    #print "cvorj", u.cvorj
                    cw_ruta.append(u.cvorj)
                    #print "1", cw_ruta
                    break
            elif u.cvorj==cw_ruta[0]:
                if u.cvori not in cw_ruta:
                    #print "cvori", u.cvori
                    #print "dadada", [u.cvori],cw_ruta
                    cw_ruta=[u.cvori] + cw_ruta
                    #print "2", cw_ruta
                    break
                #else: ustede.remove(u)
                #print "obrisao", u ne moye jer se pomjeri sledeca usteda i onda je preskoci
    cw_ruta=[prvic] + cw_ruta + [1]
    cw_duzina=izracunaj_duzinu(cw_ruta)
    #print "cw", cw_ruta, cw_duzina
            
    global solution, t

    def twoOptSwap_saduzinom(fields, i, k):
        newFields = fields[:]
        #print i,k,fields[i],fields[k] 
        #print "fields1:", newFields[i-1:k+2]
        duzina1=izracunaj_duzinu (newFields[i-1:k+2])
        # ne ukljucuje i broj na poziciji k+2, i-1 kaze da uzima od jednog broja prije i, a k+2 do jednog broja poslije k
        newFields[i:k+1] = fields[k:i-1:-1]
        # prvi dio ide od i do k, a drugi od k do i. stavljeno je i-1 da bi upalo i
        #ovo -1 znaci da se broj unazad, pa onda mo�e da se stavi npr a[6:3:-1], a ne bi moglo samo a[6:3]
        #primjer a=[1,2,3,4,5,6,7,8], i=3, i=6, dobije se a=[1, 2, 3, 7, 6, 5, 4, 8]
        #print "fields2:", newFields[i-1:k+2]
        duzina2=izracunaj_duzinu (newFields[i-1:k+2])
        #print "duzina2-duzina1", duzina2, duzina1
        razlika=duzina1-duzina2
        return newFields,razlika


    solution=Route(1,cw_ruta,cw_duzina)
    
    #print "solution", solution, solution.duzina
    #print "kw", cw_ruta,cw_duzina
    sf=cw_ruta
    sf_duzina=cw_duzina
    rute_razvoza1[tuple(sf)]=[round (sf_duzina,2)]
    #rute_razvoza[tuple(cw_ruta)]=[round (cw_duzina,2),2]

    end_raz=max(5,round((broj_klijenata)/4.5))
    brojac_raz=end_raz-3
    brcv_uruti= len(cw_ruta)
    #print "brcv_uruti", brcv_uruti, rute_razvoza1

    #t =  [[0 for x in range(brcv_uruti)] for y in range(brcv_uruti)] #vrijednost tabua
    tlong =  [[0 for x in range(brcv_uruti+1)] for y in range(brcv_uruti+1)]

    def tabu_deo (fields, duzina, zadato, iterator, t1, t2, rute_razvoza1,  tlong1):
            global solution,currentbestRoute,t
            najkraca_dosad=zadato
            najrazlika=0
            steta=zadato
            provera=False
            brojrute=-1;iterator=iterator
            dobit=0
            
            for i in range(0, brcv_uruti-1):
                for k in range(i + 2, brcv_uruti-1):
                    brojrute = brojrute+1
                    #print "0nulti", i+1,k, fields [i+1], fields[k]
                    #gain = c(i, i+1) + c(k, k+1) - c(i, k) - c(i+1, k+1)
                    #gain=d[fields[i]][fields[i+1]]+d[fields[k]][fields[k+1]]-d[fields[i]][fields[k]]-d[fields[i+1]][fields[k+1]]
                    tos, razlika_nova=twoOptSwap_saduzinom(fields, i+1, k)
                    #tos=twoOptSwap(fields, i+1, k)
                    #tos5=tos[:-1]
                    ##razlika=(duzina-gain)- solution.duzina
                    razlika=(duzina-razlika_nova)- solution.duzina
                    if razlika <najrazlika:
                        #zbog ovog nekad nema stos jer budu sve zamijene zauzete pa treba da se izvrti dok se ne oslobodi.
                        if tuple(tos) not in rute_razvoza1:
                                #print razlika,duzina,gain,solution.duzina
                                #dobit=gain
                                provera=True
                                tosnaj=tos
                                inaj=fields[i+1]
                                knaj=fields[k]
                                #najduzina=duzina-gain
                                najduzina=duzina-razlika_nova
                                najrazlika=razlika
                                #print inaj, knaj
                                
                    elif provera==False:
                      if razlika + tlong[fields[i+1]][fields[k]] < steta:
                        if t[fields[i+1]][fields[k]]==0:
                        #print "taboo", fields[i+1],fields[k],t[fields[i+1]][fields[k]]
                        #if t[fields[i+1]][fields[k]]==0:
                            #print "tos ",duzina-razlika_nova,tos,rute_razvoza1
                            if tuple(tos) not in rute_razvoza1:
                                
                                steta=razlika + tlong[fields[i+1]][fields[k]]
                                sinaj=fields[i+1]
                                sknaj=fields[k]
                                stosnaj=tos
                                #print sinaj, sknaj, razlika,tlong[fields[i+1]][fields[k]],steta
                                         
                                                                                                                                                                                   
            if provera==True:
                    #print "provera je True. Zamijeni:", inaj, knaj
                    #print solution.duzina, dobit, najduzina
                    currentbestRoute=Route(brojrute,tosnaj,najduzina)
                    solution=currentbestRoute
                    #print "1prvi", len(fields),brcv_uruti
                    for i in range(1, len(fields)):
                        for k in range(1, len(fields)):
                                  if t[fields[i]][fields[k]]>0:
                                        t[fields[i]][fields[k]]=t[fields[i]][fields[k]]-1
                                        #t[fields[k]][fields[i]]=t[fields[k]][fields[i]]-1

                    tlong[knaj][inaj]+=5000;tlong[inaj][knaj]+=5000
                    t[knaj][inaj]=t1 ;t[inaj][knaj]=t2
                     
                    iterator=iterator+1
                    tabu_deo (solution.fields,solution.duzina, 50000000000000, iterator, t1, t2, rute_razvoza1, tlong1)
                    
            elif iterator<broj_klijenata:
                    
                    nova_duzina=solution.duzina+steta
                    currentbestRoute=Route(brojrute,stosnaj,nova_duzina)
                    #print "2drugi", len(fields),brcv_uruti
                    for i in range(1, len(fields)):
                        for k in range(1, len(fields)):
                            if t[fields[i]][fields[k]]>0:
                                t[fields[i]][fields[k]]=t[fields[i]][fields[k]]-1
                                #t[fields[k]][fields[i]]=t[fields[k]][fields[i]]-1
                                    
                    #print iterator, "nema provere,zamijeni", sinaj, sknaj,tlong[sknaj][sinaj]
                    tlong[sknaj][sinaj]+=5000;tlong[sinaj][sknaj]+=5000
                    t[sknaj][sinaj]=t1 ;t[sinaj][sknaj]=t2
                    

                    iterator=iterator+1
                    tabu_deo (currentbestRoute.fields,currentbestRoute.duzina, 50000000000000, iterator, t1, t2, rute_razvoza1, tlong1)
 
    while not brojac_raz>end_raz:
        try:
                t =  [[0 for x in range(brcv_uruti+1)] for y in range(brcv_uruti+1)]
                tlong1=tlong
                tabu_deo (cw_ruta,cw_duzina, 500000000000000000,0,brojac_raz,brojac_raz, rute_razvoza1,tlong1)
                #print "novi taboooooo:", brojac_raz,solution.duzina
        except Exception as e:
                print(e)
                print ("except razvozne")
                #tabu_deo (solution.fields,solution.duzina, 50000000000000, 0, brojac_raz, brojac_raz, rute_razvoza, tlong1)

        
    
        
        if solution.duzina<sf_duzina:
            sf_duzina=solution.duzina
            sf=solution.fields
            #print "dole",sf, sf_duzina, solution.fields
            rute_razvoza1[tuple(sf)]=[round (sf_duzina,2)]
            #print rute_razvoza1
        

        brojac_raz+=1
        solution=Route(1,cw_ruta,cw_duzina)
        #print "dole cw_ruta", len(cw_ruta), cw_ruta
    
            
    rute_razvoza[tuple(sf[:-1])]=[round (sf_duzina,2)]
    #print "rute povrata", rute_razvoza

        
        
    return (rute_razvoza)

"""
    f = open('D:\Google Drive\sf', 'w')
    simplejson.dump(sf, f)
    f.close()
"""




    
            
             
