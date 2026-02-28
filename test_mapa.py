"""
Brz test za generisanje mape sa rutama - bez većine modula
"""
import folium
from openpyxl import load_workbook
import itertools
import os

# Učitaj koordinate iz Excel-a
print("📂 Učitavanje koordinata...")
wb = load_workbook(filename='data/hello.xlsx')
ws = wb.active
brk = ws.max_row

vector = [0]  # index 0 je dummy
for i in range(1, brk + 1):
    lat = float(ws.cell(row=i, column=1).value)
    lon = float(ws.cell(row=i, column=2).value)
    vector.append([lat, lon])

print(f"✅ Učitano {len(vector)-1} čvorova")

# Primer rute (ako nema podataka, koristi test rutu)
# Struktura: {ruta_id: ([nod list], other_data, povratni_index)}
test_rute = {
    'route_1': ([1, 2, 3, 4, 5, 1], None, 4),
    'route_2': ([1, 6, 7, 8, 9, 1], None, 4),
}

print("\n🗺️  Generisanje mape...")

# Kreiraj mapu
avg_lat = sum(vector[i][0] for i in range(1, len(vector))) / (len(vector) - 1)
avg_lon = sum(vector[i][1] for i in range(1, len(vector))) / (len(vector) - 1)

mapa = folium.Map(
    location=[avg_lat, avg_lon],
    zoom_start=13,
    tiles='OpenStreetMap'
)

# Boje za rute
color_list = ["green", "red", "blue", "purple", "orange", "darkred"]
color_cycle = itertools.cycle(color_list)

# Dodaj rute na mapu
for route_id, (nodes, _, povratni_idx) in test_rute.items():
    boja = next(color_cycle)
    
    # Ekstrakcija koordinata
    coords = []
    for node_id in nodes:
        if 0 < node_id < len(vector):
            coords.append([vector[node_id][0], vector[node_id][1]])
    
    if len(coords) == 0:
        continue
    
    # Crtaj liniju
    folium.PolyLine(
        coords,
        color=boja,
        weight=3,
        opacity=0.8,
        popup=f'Ruta {route_id}'
    ).add_to(mapa)
    
    # Dodaj markere
    for idx, (lat, lon) in enumerate(coords):
        node_num = nodes[idx]
        
        if idx < povratni_idx:
            marker_color = 'blue'
            icon_txt = '📦'
        else:
            marker_color = 'red'
            icon_txt = '↩️'
        
        folium.CircleMarker(
            location=[lat, lon],
            radius=6,
            popup=f'Čvor {node_num}',
            color=marker_color,
            fill=True,
            fillColor=marker_color,
            fillOpacity=0.7
        ).add_to(mapa)

# Spremi mapu
output_file = 'data/rute_mapa.html'
mapa.save(output_file)

print(f"✅ Mapa sprema na: {output_file}")
print(f"📊 Broj ruta: {len(test_rute)}")
print("\n💻 Otvorite fajl u pregledniku da vidite mapu!")
print(f"File: file://{os.path.abspath(output_file)}")
