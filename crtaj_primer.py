"""
Brza skripta za prikaz ruta - test fajl
"""
from crtaj_rute import crtaj
import ulazniPodaciPrimjer
from openpyxl import load_workbook

# Učitajte podatke
wb = load_workbook(filename='data/hello.xlsx')
ws = wb.active
brk = ws.max_row  # broj čvorova

# Napravite vector koordinata
vector_opsti = [0 for x in range(brk + 1)]
for i in range(1, brk + 1):
    vector_opsti[i] = [float(ws.cell(row=i, column=1).value), 
                        float(ws.cell(row=i, column=2).value)]

# Dobijte rute iz vašeg algoritma (primer sa Clark-Wright ili LP)
# frute_iz_res = vase_rute_iz_algoritma

# NACRTAJ RUTE
# crtaj(vector_opsti, frute_iz_res)

print("Sprema te je za crtanje ruta!")
print("Zamените 'frute_iz_res' sa vašim rutama i pokrenite crtaj()")
