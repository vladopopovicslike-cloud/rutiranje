from openpyxl import load_workbook
from collections import defaultdict
import xlsxwriter 
import pickle, random
from random import choice


#treba sabrati sva dicta, al problemje sto ne smije biti duplicat kljuc, treba provjeriti al mislim da iz ove opste uzima samo rastojanje  ivrijeme izmedju. ne i osobine a njih i ne treba

def izracunaj_matricu (ws, ws1, ws2, brk, minKoleta, randomklijenti, kolicine):
        
        randomklijenti=[1]+randomklijenti
        brojCvorova=len(randomklijenti)
        kolicine[1]=0

        #print randomklijenti_poz

        #d = [[0 for x in range(brk+1)] for y in range(brk+1)]
        #durations = [[0 for x in range(brk+1)] for y in range(brk+1)]

        d = [[0 for x in range(brojCvorova+1)] for y in range(brojCvorova+1)]
        durations = [[0 for x in range(brojCvorova+1)] for y in range(brojCvorova+1)]

        vector=[0 for x in range(brk+1)]
        br_koleta=0; uk_masa=0; osobine=defaultdict(list);broj_klijenata=0;broj_klijenata_poz=0;broj_klijenata_neg=0
        for i in randomklijenti:
                #if int((ws.cell(row = i, column = 3).value))>=minKoleta: ovo treba ako se za sve klijente daje odjednom i za poz i neg
                # ili kolicine [i]
                        broj_klijenata=broj_klijenata+1; j1=0
                        osobine[broj_klijenata].append(kolicine[i])
                        osobine[broj_klijenata].append(int((ws.cell(row = i, column = 4).value)))
                        osobine[broj_klijenata].append(int((ws.cell(row = i, column = 5).value)))
                        osobine[broj_klijenata].append(int((ws.cell(row = i, column = 6).value)))
                        osobine[broj_klijenata].append(i)
                

                        #vector[broj_klijenata]=[float(ws.cell(row = i, column = 1).value),float(ws.cell(row = i, column = 2).value)]
                        br_koleta=br_koleta + abs(float(ws.cell(row = i, column = 3).value))
                        uk_masa=uk_masa + float(ws.cell(row = i, column = 4).value)
                      
                        for j in randomklijenti:
                                #if int((ws.cell(row = j, column = 3).value))>=minKoleta:
                                        j1=j1+1
                                        d[broj_klijenata][j1]=float(ws1.cell(row = i, column = j).value)
                                        durations[broj_klijenata][j1]=float(ws2.cell(row = i, column = j).value)
                                        #print i,j, broj_klijenata,j1, float(ws1.cell(row = i, column = j).value)
        #broj_klijenata_poz=broj_klijenata-1                
        return (d, durations, osobine, broj_klijenata, vector, br_koleta, uk_masa)

def izracunaj_matricu_neg (ws, ws1, ws2, brk, minKoleta, randomklijenti, kolicine):
        
        randomklijenti=[1]+randomklijenti
        brojCvorova=len(randomklijenti)
        kolicine[1]=0

        #print randomklijenti_poz

        #d= [[0 for x in range(brk+1)] for y in range(brk+1)]
        #durations = [[0 for x in range(brk+1)] for y in range(brk+1)]

        d = [[0 for x in range(brojCvorova+1)] for y in range(brojCvorova+1)]
        durations = [[0 for x in range(brojCvorova+1)] for y in range(brojCvorova+1)]

        vector=[0 for x in range(brk+1)]
        br_koleta=0; uk_masa=0; osobine=defaultdict(list);broj_klijenata=0;broj_klijenata_poz=0;broj_klijenata_neg=0
        for i in randomklijenti:
                #if int((ws.cell(row = i, column = 3).value))>=minKoleta: ovo treba ako se za sve klijente daje odjednom i za poz i neg
                # ili kolicine [i]
                        #print broj_klijenata, i, kolicine[i]
                        broj_klijenata=broj_klijenata+1; j1=0
                        osobine[broj_klijenata].append(kolicine[i])
                        osobine[broj_klijenata].append(int((ws.cell(row = abs(i), column = 4).value)))
                        osobine[broj_klijenata].append(int((ws.cell(row = abs(i), column = 5).value)))
                        osobine[broj_klijenata].append(int((ws.cell(row = abs(i), column = 6).value)))
                        osobine[broj_klijenata].append(i)
                

                        #vector[broj_klijenata]=[float(ws.cell(row = abs(i), column = 1).value),float(ws.cell(row = abs(i), column = 2).value)]
                        br_koleta=br_koleta + abs(float(ws.cell(row = abs(i), column = 3).value))
                        uk_masa=uk_masa + float(ws.cell(row = abs(i), column = 4).value)
                      
                        for j in randomklijenti:
                                #if int((ws.cell(row = j, column = 3).value))>=minKoleta:
                                        j1=j1+1
                                        d[broj_klijenata][j1]=float(ws1.cell(row = abs(i), column = abs(j)).value)
                                        durations[broj_klijenata][j1]=float(ws2.cell(row = abs(i), column = abs(j)).value)
                                        #print i,j, broj_klijenata,j1, float(ws1.cell(row = i, column = j).value)
        #broj_klijenata_poz=broj_klijenata-1                
        return (d, durations, osobine, broj_klijenata, vector, br_koleta, uk_masa)


def izracunaj_matricu_opstu (ws, ws1, ws2,brk, randomklijenti_poz, kolicine_poz, randomklijenti_neg, kolicine_neg):
        
        randomklijenti_poz=[1]+randomklijenti_poz
        #brojCvorova_poz=len(randomklijenti_poz)
        kolicine_poz[1]=0

        randomklijenti_neg=[1]+randomklijenti_neg
        #brojCvorova_neg=len(randomklijenti_neg)
        kolicine_neg[1]=0

        #brojCvorova=brojCvorova_poz+brojCvorova_neg

        d = [[0 for x in range(brk+1)] for y in range(brk+1)]
        durations = [[0 for x in range(brk+1)] for y in range(brk+1)]
        osobine=defaultdict(list)

        #d={};durations={}

        broj_klijenata=0
        for i in randomklijenti_poz:
                broj_klijenata=broj_klijenata+1; j1=0
                osobine["poz", i].append(kolicine_poz[i])
                osobine["poz", i].append(int((ws.cell(row = i, column = 4).value)))
                osobine["poz", i].append(int((ws.cell(row = i, column = 5).value)))
                osobine["poz", i].append(int((ws.cell(row = i, column = 6).value)))
                                      
                for j in randomklijenti_poz:
                        d[i][j]=float(ws1.cell(row = i, column = j).value)
                        durations[i][j]=float(ws2.cell(row = i, column = j).value)
                for j in randomklijenti_neg:
                        d[i][abs(j)]=float(ws1.cell(row = i, column = abs(j)).value)
                        durations[i][abs(j)]=float(ws2.cell(row = i, column = abs(j)).value)
                        

        broj_klijenata=0
        for i in randomklijenti_neg:
                broj_klijenata=broj_klijenata+1; j1=0
                osobine["neg", i].append(kolicine_neg[i])
                osobine["neg", i].append(int((ws.cell(row = abs(i), column = 4).value)))
                osobine["neg", i].append(int((ws.cell(row = abs(i), column = 5).value)))
                osobine["neg", i].append(int((ws.cell(row = abs(i), column = 6).value)))
                                      
                for j in randomklijenti_neg:
                        d[abs(i)][abs(j)]=float(ws1.cell(row = abs(i), column = abs(j)).value)
                        durations[abs(i)][abs(j)]=float(ws2.cell(row = abs(i), column = abs(j)).value)

                for j in randomklijenti_poz:
                        d[abs(i)][j]=float(ws1.cell(row = abs(i), column = j).value)
                        durations[abs(i)][j]=float(ws2.cell(row = abs(i), column = j).value)
                        
        return (d, durations, osobine)



def novo_stanje(period,randomklijenti_poz,kolicine_poz, randomklijenti_neg,kolicine_neg,randomklijenti_svi,randomklijenti_neg_sve_liste):
        if period==0:
                #prva sema novo stanje (jutro)
                kolicine_poz=dict.fromkeys(randomklijenti_poz[0:20], 1)
                kolicine_poz.update(dict.fromkeys(randomklijenti_poz[20:30], 2))
                kolicine_poz.update(dict.fromkeys(randomklijenti_poz[30:40], 3))
                kolicine_poz.update(dict.fromkeys(randomklijenti_poz[40:50], 2))
                kolicine_poz.update(dict.fromkeys(randomklijenti_poz[50:60], 3))
                kolicine_poz.update(dict.fromkeys(randomklijenti_poz[60:70], 2))

                #print randomklijenti_neg_sve_liste[0]
                randomklijenti_neg=randomklijenti_neg_sve_liste[0]
                
                kolicine_neg=dict.fromkeys(randomklijenti_neg[0:12], 1)
                kolicine_neg.update(dict.fromkeys(randomklijenti_neg[12:15], 2))

                randomklijenti_svi = randomklijenti_poz + randomklijenti_neg
                #kolicine_sve=dict.fromkeys(randomklijenti_svi, -1)
                        
                sema=[randomklijenti_poz,kolicine_poz, randomklijenti_neg,kolicine_neg,randomklijenti_svi]
                
        if period==1:
                #druga sema novo stanje (podne)
                randomklijenti_poz2=randomklijenti_poz[20:70]
                kolicine_poz=dict.fromkeys(randomklijenti_poz2[0:10], 2)
                kolicine_poz.update(dict.fromkeys(randomklijenti_poz2[10:50], 3))

                #print randomklijenti_neg_sve_liste[1]
                randomklijenti_neg=randomklijenti_neg_sve_liste[1][5:15]
                
                kolicine_neg=dict.fromkeys(randomklijenti_neg[0:5], 2)
                kolicine_neg.update(dict.fromkeys(randomklijenti_neg[5:7], 1))
                kolicine_neg.update(dict.fromkeys(randomklijenti_neg[7:10], 2))

                randomklijenti_svi = randomklijenti_poz2 + randomklijenti_neg
                #kolicine_sve=dict.fromkeys(randomklijenti_svi, -1)
                        
                sema=[randomklijenti_poz2,kolicine_poz, randomklijenti_neg,kolicine_neg,randomklijenti_svi]

        if period==2:
                #treca sema novo stanje (vece)
                randomklijenti_poz3=randomklijenti_poz[0:30]+randomklijenti_poz[40:70]
                # u sl redu je stajalo 2 na kraju
                kolicine_poz=dict.fromkeys(randomklijenti_poz3[0:30], 1)
                kolicine_poz.update(dict.fromkeys(randomklijenti_poz3[30:40], 1))
                kolicine_poz.update(dict.fromkeys(randomklijenti_poz3[40:50], 3))
                kolicine_poz.update(dict.fromkeys(randomklijenti_poz3[50:60], 4))

                #print randomklijenti_neg_sve_liste[2]
                randomklijenti_neg=randomklijenti_neg_sve_liste[2][0:5]+randomklijenti_neg_sve_liste[2][10:15]
                
                kolicine_neg=dict.fromkeys(randomklijenti_neg[0:5], 2)
                kolicine_neg.update(dict.fromkeys(randomklijenti_neg[5:7], 1))
                kolicine_neg.update(dict.fromkeys(randomklijenti_neg[7:10], 2))

                randomklijenti_svi = randomklijenti_poz3 + randomklijenti_neg
                #kolicine_sve=dict.fromkeys(randomklijenti_svi, -1)
                        
                sema=[randomklijenti_poz3,kolicine_poz, randomklijenti_neg,kolicine_neg,randomklijenti_svi]

        return (sema)


def generisanje_sema(broj_sema):
    sve_seme=[]    
    for i in range(broj_sema):
        """randomklijenti_poz = random.sample(range(2, 201), 45)
        randomklijenti_neg = random.sample(randomklijenti_poz, 10)
        for i in range (10):
                randomklijenti_neg.append(choice([i for i in range(2,201) if i not in randomklijenti_neg]))
	
        #randomklijenti_neg = random.sample(range(2, 201), 30)
        randomklijenti_neg =[element*(-1) for element in randomklijenti_neg]
        """
        randomklijenti_neg_sve_liste=[]
        #randomklijenti_poz = [i for i in range (2,72)]
        randomklijenti_poz = random.sample(range(2, 201), 70)             

        #randomklijenti_poz=[3, 183, 119, 38, 173, 134, 156, 97, 88, 145, 170, 69, 4, 192, 195, 2, 59, 52, 84, 164, 66, 115, 37, 35, 92, 8, 123, 131, 27, 167, 21, 136, 146, 24, 122, 187, 194, 178, 64, 28, 36, 77, 200, 158, 108, 95, 182, 75, 171, 120, 15, 175, 7, 114, 65, 193, 111, 166, 83, 184, 197, 61, 198, 93, 130, 110, 149, 153, 106, 142]
        #randomklijenti_poz=randomklijenti_poz[0:30]+randomklijenti_poz[40:70]
        
        kolicine_poz=dict.fromkeys(randomklijenti_poz[0:20], 1)
        kolicine_poz.update(dict.fromkeys(randomklijenti_poz[20:50], 2))
        kolicine_poz.update(dict.fromkeys(randomklijenti_poz[50:70], 3))
        #kolicine_poz.update(dict.fromkeys(randomklijenti_poz[50:60], 4))

        randomklijenti_neg_svi=[]
        for period in range(3):
                sema=[randomklijenti_poz,kolicine_poz]
                #print sema
                randomklijenti_neg=[]
                for i in range (15):
                        randomklijenti_neg.append(choice([i for i in randomklijenti_poz if i not in randomklijenti_neg]))
                        randomklijenti_neg_svi=randomklijenti_neg_svi+[randomklijenti_neg[-1]]

                randomklijenti_neg =[element*(-1) for element in randomklijenti_neg]
                #print randomklijenti_neg
                #print
                randomklijenti_neg_sve_liste.append(randomklijenti_neg)
                
                kolicine_neg=dict.fromkeys(randomklijenti_neg[0:12], 1)
                kolicine_neg.update(dict.fromkeys(randomklijenti_neg[12:15], 2))
                #kolicine_neg.update(dict.fromkeys(randomklijenti_neg[7:10], 2))

                randomklijenti_svi = randomklijenti_poz + randomklijenti_neg
                kolicine_sve=dict.fromkeys(randomklijenti_svi, -1)
                
                sema=sema+[randomklijenti_neg,kolicine_neg,randomklijenti_svi]
                sve_seme.append(sema)

                #ovo isključiš i uključi ispod i dobijes prvo sva stara stanja pa sva nova
                sema=novo_stanje(period,randomklijenti_poz,kolicine_poz, randomklijenti_neg,kolicine_neg,randomklijenti_svi,randomklijenti_neg_sve_liste)
                sve_seme.append(sema)
                
        
        """#prva sema novo stanje (jutro)
        kolicine_poz=dict.fromkeys(randomklijenti_poz[0:20], 1)
        kolicine_poz.update(dict.fromkeys(randomklijenti_poz[20:30], 2))
        kolicine_poz.update(dict.fromkeys(randomklijenti_poz[30:40], 3))
        kolicine_poz.update(dict.fromkeys(randomklijenti_poz[40:50], 2))
        kolicine_poz.update(dict.fromkeys(randomklijenti_poz[50:60], 3))
        kolicine_poz.update(dict.fromkeys(randomklijenti_poz[60:70], 2))

        #print randomklijenti_neg_sve_liste[0]
        randomklijenti_neg=randomklijenti_neg_sve_liste[0]
        
        kolicine_neg=dict.fromkeys(randomklijenti_neg[0:12], 1)
        kolicine_neg.update(dict.fromkeys(randomklijenti_neg[12:15], 2))

        randomklijenti_svi = randomklijenti_poz + randomklijenti_neg
        kolicine_sve=dict.fromkeys(randomklijenti_svi, -1)
                
        sema=[randomklijenti_poz,kolicine_poz, randomklijenti_neg,kolicine_neg,randomklijenti_svi]
        sve_seme.append(sema)

        #druga sema novo stanje (podne)
        randomklijenti_poz2=randomklijenti_poz[20:70]
        kolicine_poz=dict.fromkeys(randomklijenti_poz2[0:10], 2)
        kolicine_poz.update(dict.fromkeys(randomklijenti_poz2[10:50], 3))

        #print randomklijenti_neg_sve_liste[1]
        randomklijenti_neg=randomklijenti_neg_sve_liste[1][5:15]
        
        kolicine_neg=dict.fromkeys(randomklijenti_neg[0:5], 2)
        kolicine_neg.update(dict.fromkeys(randomklijenti_neg[5:7], 1))
        kolicine_neg.update(dict.fromkeys(randomklijenti_neg[7:10], 2))

        randomklijenti_svi = randomklijenti_poz2 + randomklijenti_neg
        kolicine_sve=dict.fromkeys(randomklijenti_svi, -1)
                
        sema=[randomklijenti_poz2,kolicine_poz, randomklijenti_neg,kolicine_neg,randomklijenti_svi]
        sve_seme.append(sema)

        #treca sema novo stanje (vece)
        randomklijenti_poz3=randomklijenti_poz[0:30]+randomklijenti_poz[40:70]
        
        kolicine_poz=dict.fromkeys(randomklijenti_poz3[0:30], 2)
        kolicine_poz.update(dict.fromkeys(randomklijenti_poz3[30:40], 1))
        kolicine_poz.update(dict.fromkeys(randomklijenti_poz3[40:50], 3))
        kolicine_poz.update(dict.fromkeys(randomklijenti_poz3[50:60], 4))

        #print randomklijenti_neg_sve_liste[2]
        randomklijenti_neg=randomklijenti_neg_sve_liste[2][0:5]+randomklijenti_neg_sve_liste[2][10:15]
        
        kolicine_neg=dict.fromkeys(randomklijenti_neg[0:5], 2)
        kolicine_neg.update(dict.fromkeys(randomklijenti_neg[5:7], 1))
        kolicine_neg.update(dict.fromkeys(randomklijenti_neg[7:10], 2))

        randomklijenti_svi = randomklijenti_poz3 + randomklijenti_neg
        kolicine_sve=dict.fromkeys(randomklijenti_svi, -1)
                
        sema=[randomklijenti_poz3,kolicine_poz, randomklijenti_neg,kolicine_neg,randomklijenti_svi]
        sve_seme.append(sema)"""


    return (sve_seme)
        
